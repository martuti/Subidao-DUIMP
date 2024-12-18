import json
import pandas as pd
from flask import Flask, request, jsonify, render_template
import os
from datetime import datetime
import time
from openpyxl import load_workbook
import subprocess
import signal


 
app = Flask(__name__)

atributos = '\\ATRIBUTOS_POR_NCM_Produção.json'
atributos = '\\ATRIBUTOS_POR_NCM_Validação.json'

path = "C:\\Users\\amartu01\\Documents\\DUIMP - Cherps\\Subidão DUIMP\\Plan DUIMP Project"

#Verifica se a planilha Lista de Itens DUIMP está aberta ou fechada
planilha_path = path + "\\Lista de Itens DUIMP.xlsx"

while True:
    try:
        # Tenta abrir o arquivo
        df = pd.read_excel(
            planilha_path,
            sheet_name="Sheet1",
            usecols=["Produto", "NCM", "Forn 1", "País", "Des Fornec 1", "codigo", "Descricao Sistema"],
            dtype=str
        ).dropna()
        print("Arquivo carregado com sucesso!")
        break  # Sai do laço se o arquivo for carregado com sucesso
    except PermissionError:
        print(f"Erro: O arquivo '{planilha_path}' está aberto. Por favor, feche o arquivo e pressione Enter para continuar.")
        input("Pressione Enter depois de fechar o arquivo...")
        time.sleep(1)  # Aguarda um segundo antes de tentar novamente


output_dir = path + "\\Planilha Catálogo"
if not os.path.isdir(output_dir):
    raise FileNotFoundError(f"A pasta fornecida não existe: {output_dir}")

# Carregar o JSON
try:
    with open(path + atributos, 'r', encoding='utf-8') as file:
        dados = json.load(file)
except FileNotFoundError as e:
    raise FileNotFoundError(f"Arquivo JSON não encontrado: {e}")

json_path = path + atributos

# Função para carregar o JSON globalmente
def carregar_json():
    try:
        with open(json_path, 'r', encoding='utf-8') as file:
            return json.load(file)
    except FileNotFoundError:
        raise FileNotFoundError("Arquivo JSON não encontrado.")

# Variável global para armazenar o JSON carregado
dados_json = carregar_json()

# Carregar a planilha
planilha_path = path + '\\Lista de Itens DUIMP.xlsx'
try:
    df = pd.read_excel(planilha_path, sheet_name="Sheet1", usecols=["Produto", "NCM", "Forn 1", "País", "Des Fornec 1", "codigo", "Descricao Sistema"], dtype=str).dropna()
except KeyError as e:
    raise ValueError(f"A coluna esperada não foi encontrada na planilha. Verifique o nome da coluna: {e}")
except FileNotFoundError as e:
    raise FileNotFoundError(f"Arquivo da planilha não encontrado: {e}")

@app.route('/get_pn_options', methods=['GET'])
def get_pn_options():
    """
    Retorna os valores únicos de Produto (PN) e suas NCMs para preencher o datalist.
    """
    pn_options = df.to_dict('records')  # [{'Produto': '...', 'NCM': '...'}, ...]
    return jsonify(pn_options)

def buscar_detalhes_ncm(codigo_ncm):
    """
    Busca os detalhes da NCM no JSON.
    """
    codigo_ncm_normalizado = codigo_ncm.replace('.', '')
    ncm = next(
        (item for item in dados["listaNcm"] if item["codigoNcm"].replace('.', '') == codigo_ncm_normalizado), 
        None
    )
    if not ncm:
        return None

    atributos = []
    for atributo in ncm.get("listaAtributos", []):
        if atributo.get("modalidade") == "Importação":
            detalhes = next((det for det in dados["detalhesAtributos"] if det["codigo"] == atributo["codigo"]), None)
            if detalhes:
                atributos.append(detalhes)
    return atributos

def validar_condicionados(atributo_codigo, resposta):
    """
    Valida os atributos condicionados com base na resposta fornecida.
    """
    atributo = next((attr for attr in dados["detalhesAtributos"] if attr["codigo"] == atributo_codigo), None)
    if not atributo or not atributo.get("condicionados"):
        return []

    condicionados = []
    for condicionado in atributo["condicionados"]:
        condicao = condicionado["condicao"]
        if condicao["operador"] == "==" and resposta == condicao["valor"]:
            condicionados.append(condicionado["atributo"])
            # Chamado recursivamente para buscar outros condicionados dentro dos próprios atributos condicionados
            condicionados.extend(validar_condicionados(condicionado["atributo"]["codigo"], resposta))
    return condicionados

@app.route('/')
def index():
    return render_template('formulario.html')

@app.route('/buscar_ncm', methods=['POST'])
def buscar_ncm():
    codigo_ncm = request.json.get("ncm")
    if not codigo_ncm:
        return jsonify({"error": "Código NCM não fornecido"}), 400

    atributos = buscar_detalhes_ncm(codigo_ncm)
    if not atributos:
        return jsonify({"error": "NCM não encontrada"}), 404

    return jsonify(atributos)

@app.route('/validar_condicao', methods=['POST'])
def validar_condicao():
    atributo_codigo = request.json.get("atributoCodigo")
    resposta = request.json.get("resposta")

    if not atributo_codigo or resposta is None:
        return jsonify({"error": "Dados insuficientes para validação"}), 400

    condicionados = validar_condicionados(atributo_codigo, resposta)
    return jsonify(condicionados)

def consolidar_atributos(lista_ncm, detalhes_atributos):
    obrigatorios = {atributo["codigo"]: atributo["obrigatorio"] for atributo in lista_ncm}
    for atributo in detalhes_atributos:
        atributo["obrigatorio"] = obrigatorios.get(atributo["codigo"], atributo.get("obrigatorio", False))
    return detalhes_atributos

@app.route('/salvar_respostas', methods=['POST'])
def salvar_respostas():
    try:
        data = request.json
        if not data:
            return jsonify({"error": "Nenhum dado recebido"}), 400

        # Captura campos principais
        campos_principais = {
            "ncm": data.get("ncm", "").strip(),
            "pn": data.get("pn", "").strip(),
            "fornecedor": data.get("fornecedor", "").strip(),
            "nome_fornecedor": data.get("nomeFornecedor", "").strip(),
            "pais_fornecedor": data.get("paisFornecedor", "").strip(),
            "codigo_fornecedor_duimp": data.get("codigoFornecedorDuimp", "").strip(),
            "denominacao": data.get("denominacao", "").strip(),
            "descricao_detalhada": (
                f"{data.get('denominacao', '').strip()}, "
                f"O que é: {data.get('oqueE', '').strip()}, "
                f"Função: {data.get('funcao', '').strip()}, "
                f"Aplicação: {data.get('aplicacao', '').strip()}, "
                f"Material Constitutivo: {data.get('materialConstitutivo', '').strip()}, "
                f"Marca: {data.get('marca', '').strip()}, "
                f"Modelo: {data.get('modelo', '').strip()}"
            )
        }

        respostas = data.get("respostas", [])
        if not respostas:
            return jsonify({"error": "Nenhuma resposta foi enviada"}), 400

        registros = []

        def buscar_detalhes(codigo, atributos):
            for atributo in atributos:
                if atributo["codigo"] == codigo:
                    return atributo
                for condicionado in atributo.get("condicionados", []):
                    detalhes = buscar_detalhes(codigo, [condicionado["atributo"]])
                    if detalhes:
                        return detalhes
            return {}

        def processar_resposta(resposta, atributos, nivel=0):
            atributo_detalhes = buscar_detalhes(resposta["name"], atributos)
            nome_apresentacao = atributo_detalhes.get("nomeApresentacao", "Não encontrado")
            forma_preenchimento = atributo_detalhes.get("formaPreenchimento", "Não encontrado")
            valor_resposta = resposta.get("value", "")

            if forma_preenchimento == "COMPOSTO":
                lista_subatributos = atributo_detalhes.get("listaSubatributos", [])
                if lista_subatributos:
                    for subatributo in lista_subatributos:
                        subresposta = {
                            "name": subatributo["codigo"],
                            "value": resposta.get("subatributos", {}).get(subatributo["codigo"], "")
                        }
                        processar_resposta(subresposta, atributos, nivel + 1)

            elif forma_preenchimento == "BOOLEANO":
                valor_resposta = "Sim" if valor_resposta == "true" else "Nao" if valor_resposta == "false" else valor_resposta
            elif forma_preenchimento == "LISTA_ESTATICA":
                dominio_item = next(
                    (dom for dom in atributo_detalhes.get("dominio", []) if dom["codigo"] == valor_resposta),
                    {}
                )
                if dominio_item:
                    valor_resposta = f"{dominio_item['codigo']} - {dominio_item['descricao']}"

            registro = {
                "Quantidade Atributos": len(respostas),
                **campos_principais,
                "Atributo": resposta.get("name", ""),
                "Nome Apresentação": nome_apresentacao,
                "Forma Preenchimento": forma_preenchimento,
                "Resposta": valor_resposta,
            }
            registros.append(registro)

        for resposta in respostas:
            processar_resposta(resposta, dados_json.get("detalhesAtributos", []))

        # Salvar registros na planilha
        planilha_path = path + r"\Planilha Catálogo\CATALOGO ROBO.xlsx"

        if not os.path.exists(planilha_path):
            return jsonify({"error": f"Arquivo não encontrado: {planilha_path}"}), 400

        try:
            wb = load_workbook(planilha_path)
            ws = wb.active

            # Encontrar a primeira linha em branco na coluna A
            primeira_linha_vazia = len(ws['A']) + 1
            for i, registro in enumerate(registros, start=primeira_linha_vazia):
                ws.cell(row=i, column=1, value=registro.get("Quantidade Atributos", ""))
                ws.cell(row=i, column=2, value=registro.get("ncm", ""))
                ws.cell(row=i, column=3, value=registro.get("pn", ""))
                ws.cell(row=i, column=4, value=registro.get("fornecedor", ""))
                ws.cell(row=i, column=5, value=registro.get("nome_fornecedor", ""))
                ws.cell(row=i, column=6, value=registro.get("pais_fornecedor", ""))
                ws.cell(row=i, column=7, value=registro.get("codigo_fornecedor_duimp", ""))
                ws.cell(row=i, column=8, value=registro.get("denominacao", ""))
                ws.cell(row=i, column=9, value=registro.get("descricao_detalhada", ""))
                ws.cell(row=i, column=10, value=registro.get("Atributo", ""))
                ws.cell(row=i, column=11, value=registro.get("Nome Apresentação", ""))
                ws.cell(row=i, column=12, value=registro.get("Forma Preenchimento", ""))
                ws.cell(row=i, column=13, value=registro.get("Resposta", ""))

            if registro.get("Forma Preenchimento") == "NUMERO_INTEIRO" and isinstance(registro.get("Resposta"), int):
                ws.cell(row=i, column=13, value=registro.get("Resposta"))
            else:
                ws.cell(row=i, column=13, value=registro.get("Resposta", ""))

            # Salvar e fechar a planilha
            wb.save(planilha_path)
            wb.close()

            # Chama o outro script Python
            
            script_path = os.path.join(r"C:\\Users\\amartu01\\Documents\\DUIMP - Cherps\\Subidão DUIMP\\Robo DUIMP", "automacao_duimp_cherps.py")
            try:
                subprocess.run(["python", script_path], check=True)
            except subprocess.CalledProcessError as e:
                return jsonify({"error": f"Erro ao executar o script: {e}"}), 500

            return jsonify({"message": "Arquivo salvo com sucesso na planilha e script executado com sucesso."})
        except Exception as e:
            return jsonify({"error": f"Erro ao salvar na planilha: {str(e)}"}), 500

    except Exception as e:
        return jsonify({"error": f"Erro ao processar os dados: {str(e)}"}), 500


@app.route('/buscar_atributo_outros', methods=['POST'])
def buscar_atributo_outros():
    data = request.json
    codigo = data.get("codigo")
    atributo_pai = data.get("atributoPai")

    if not codigo or not atributo_pai:
        return jsonify({"error": "Código ou atributo pai não fornecido"}), 400

    # Busca o atributo relacionado no JSON ou banco de dados
    atributo_outros = next(
        (attr for attr in dados_json.get("detalhesAtributos", [])
         if attr.get("codigoPai") == atributo_pai and attr.get("codigo") == codigo),
        None
    )

    if not atributo_outros:
        return jsonify({"error": "Atributo relacionado não encontrado"}), 404

    return jsonify(atributo_outros)

def depurar_atributos_nao_encontrados(respostas):
    """
    Exibe no terminal informações detalhadas dos atributos que constam como "Não encontrado".
    """
    for resposta in respostas:
        codigo = resposta.get("name", "Código não informado")
        atributo_detalhes = next(
            (attr for attr in dados_json.get("detalhesAtributos", []) if attr["codigo"] == codigo),
            {}
        )
        nome_apresentacao = atributo_detalhes.get("nomeApresentacao", "Não encontrado")
        forma_preenchimento = atributo_detalhes.get("formaPreenchimento", "Não encontrado")
        print(f"Atributo: {codigo} | Nome Apresentação: {nome_apresentacao} | Forma Preenchimento: {forma_preenchimento}")

encerrando = False  # Flag para sinalizar que a aplicação está encerrando

@app.route('/encerrar', methods=['POST'])
def encerrar():
    global encerrando
    encerrando = True
    print("Recebido comando para encerrar aplicação.")  # Log de depuração
    os.kill(os.getpid(), signal.SIGINT)  # Encerra o servidor Flask
    return jsonify({"message": "Aplicação encerrada com sucesso."})

from datetime import datetime
from openpyxl import load_workbook

def mover_item_para_feito(planilha_path, linha_item):
    """
    Move um item da aba `TBD_ListaItensDuimp` para a aba `Feito_ListaItensDuimp`
    e registra a data atual na coluna 'Cadastro Feito'.
    
    :param planilha_path: Caminho completo para a planilha.
    :param linha_item: Índice da linha a ser movida (baseado no DataFrame).
    """
    try:
        # Carregar a planilha
        wb = load_workbook(planilha_path)
        
        # Selecionar as abas
        aba_tbd = wb["Sheet1"]
        aba_feito = wb["Sheet1"]
        
        # Encontrar a primeira linha vazia na aba `Feito_ListaItensDuimp`
        primeira_linha_vazia = len(list(aba_feito.iter_rows(min_row=2, max_col=1, values_only=True))) + 2
        
        # Capturar os valores da linha a ser movida
        valores = [cell.value for cell in aba_tbd[linha_item + 2]]  # linha_item + 2 devido ao cabeçalho
        
        # Adicionar a data do dia na última coluna (Cadastro Feito)
        valores.append(datetime.now().strftime('%Y-%m-%d'))
        
        # Escrever os valores na primeira linha vazia da aba `Feito_ListaItensDuimp`
        for col_index, valor in enumerate(valores, start=1):
            aba_feito.cell(row=primeira_linha_vazia, column=col_index, value=valor)
        
        # Remover a linha original da aba `TBD_ListaItensDuimp`
        aba_tbd.delete_rows(idx=linha_item + 2, amount=1)  # linha_item + 2 devido ao cabeçalho
        
        # Salvar a planilha
        wb.save(planilha_path)
        wb.close()
        
        print(f"Linha {linha_item + 2} movida com sucesso para `Feito_ListaItensDuimp`.")
    
    except Exception as e:
        print(f"Erro ao mover item: {e}")


if __name__ == '__main__':
    app.run(debug=True)